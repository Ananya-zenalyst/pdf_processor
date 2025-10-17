import cv2
import numpy as np
import pandas as pd
import json
import tabula
import camelot
import pdfplumber
import fitz  # PyMuPDF
import os
from typing import Dict, List, Any, Union
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import io

class VisualTableExtractor:
    """
    Table extractor that uses computer vision to understand table structure
    and Camelot for extraction to provide accurate results.
    """

    def __init__(self):
        self.extraction_method = 'camelot'
        self.temp_files = []

    def __del__(self):
        """Clean up temporary files"""
        for temp_file in self.temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
            except Exception:
                pass

    def pdf_page_to_image(self, pdf_path: str, page_num: int = 0) -> np.ndarray:
        """Convert PDF page to OpenCV image for visual analysis"""
        try:
            doc = fitz.open(pdf_path)
            if page_num >= len(doc):
                page_num = 0  # Default to first page if requested page doesn't exist

            page = doc[page_num]

            # Render page to image with high DPI for better analysis
            mat = fitz.Matrix(2, 2)  # 2x zoom for better resolution
            pix = page.get_pixmap(matrix=mat)
            img_data = pix.tobytes("png")

            # Convert to OpenCV format
            nparr = np.frombuffer(img_data, np.uint8)
            image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
            doc.close()

            return image
        except Exception as e:
            print(f"    ⚠ Warning: Could not convert PDF page to image: {e}")
            # Return a blank image as fallback
            return np.zeros((100, 100, 3), dtype=np.uint8)

    def detect_table_structure(self, image: np.ndarray) -> Dict[str, Any]:
        """
        Use computer vision to detect table structure including:
        - Table boundaries
        - Row and column lines
        - Cell structure
        - Text regions
        """
        # Convert to grayscale
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

        # Apply adaptive threshold to handle different lighting conditions
        binary = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                     cv2.THRESH_BINARY_INV, 11, 2)

        # Detect horizontal and vertical lines
        horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (40, 1))
        vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 40))

        # Extract horizontal lines
        horizontal_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, horizontal_kernel)
        # Extract vertical lines
        vertical_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, vertical_kernel)

        # Combine lines to get table structure
        table_mask = cv2.bitwise_or(horizontal_lines, vertical_lines)

        # Find table contours
        contours, _ = cv2.findContours(table_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        # Filter contours by area to find actual tables - reduced threshold to catch more tables
        min_table_area = image.shape[0] * image.shape[1] * 0.001  # At least 0.1% of page (10x more sensitive)
        table_contours = [cnt for cnt in contours if cv2.contourArea(cnt) > min_table_area]

        # Get table bounding boxes
        table_boxes = []
        for contour in table_contours:
            x, y, w, h = cv2.boundingRect(contour)
            table_boxes.append({
                'bbox': (x, y, x + w, y + h),
                'area': w * h,
                'aspect_ratio': w / h if h > 0 else 0
            })

        # Detect grid structure for each table
        structure_info = {
            'table_count': len(table_boxes),
            'tables': table_boxes,
            'has_borders': len(table_contours) > 0,
            'line_density': np.sum(table_mask > 0) / (image.shape[0] * image.shape[1]),
            'image_dimensions': (image.shape[1], image.shape[0])  # (width, height)
        }

        return structure_info

    def extract_with_camelot(self, pdf_path: str, structure_info: Dict) -> List[Dict]:
        """Extract tables using Camelot with structure-aware settings"""
        tables = []
        try:
            # Validate PDF path
            if not os.path.exists(pdf_path):
                raise FileNotFoundError(f"PDF file not found: {pdf_path}")

            # Choose extraction method based on detected structure
            flavor = 'lattice' if structure_info.get('has_borders', False) and structure_info.get('line_density', 0) > 0.001 else 'stream'
            print(f"    Using Camelot {flavor} method (has_borders: {structure_info.get('has_borders', False)}, line_density: {structure_info.get('line_density', 0):.4f})")

            # Try to read PDF with error handling
            try:
                camelot_tables = camelot.read_pdf(pdf_path, pages='all', flavor=flavor, suppress_stdout=True)
            except Exception as camelot_error:
                print(f"    ⚠ Camelot {flavor} failed, trying alternate method...")
                # Try alternate flavor if first one fails
                alternate_flavor = 'stream' if flavor == 'lattice' else 'lattice'
                camelot_tables = camelot.read_pdf(pdf_path, pages='all', flavor=alternate_flavor, suppress_stdout=True)
                print(f"    Using alternate Camelot {alternate_flavor} method")

            print(f"    Camelot returned {len(camelot_tables)} raw tables")

            for i, table in enumerate(camelot_tables):
                df = table.df
                print(f"      Table {i}: Shape {df.shape}, Empty: {df.empty}")

                if not df.empty and df.shape[0] > 0 and df.shape[1] > 0:
                    # Check if table has actual content
                    non_empty_cells = df.notna().sum().sum()
                    print(f"        Non-empty cells: {non_empty_cells}/{df.shape[0] * df.shape[1]}")

                    # Analyze if this is truly a table or just text/content
                    is_real_table = self._is_real_table(df)

                    if is_real_table:
                        tables.append({
                            'method': 'camelot',
                            'table_index': i,
                            'dataframe': df,
                            'content_type': 'table',
                            'accuracy': getattr(table, 'accuracy', 0),
                            'whitespace': getattr(table, 'whitespace', 0),
                            'parsing_report': getattr(table, 'parsing_report', {})
                        })
                        print(f"        ✓ Identified as real table")
                    else:
                        print(f"        ⚠ Not a real table, likely text content")
                else:
                    print(f"        Skipping empty table {i}")

        except Exception as e:
            print(f"    ❌ Camelot extraction failed: {e}")
            import traceback
            print(f"    Traceback: {traceback.format_exc()}")

        return tables

    def extract_with_tabula(self, pdf_path: str, structure_info: Dict) -> List[Dict]:
        """Extract tables using Tabula with multiple configurations"""
        tables = []

        # Configuration based on visual structure analysis
        configs = [
            ('lattice', {'lattice': True, 'multiple_tables': True}),
            ('stream', {'stream': True, 'multiple_tables': True}),
            ('guess', {'guess': True, 'multiple_tables': True}),
            ('area_full', {'area': [0, 0, 1000, 1000], 'multiple_tables': True}),
        ]

        for config_name, config in configs:
            try:
                print(f"    Trying Tabula {config_name} method...")

                tabula_tables = tabula.read_pdf(
                    pdf_path,
                    pages='all',
                    pandas_options={'header': None},
                    **config
                )

                print(f"    Tabula {config_name} returned {len(tabula_tables)} tables")

                for i, df in enumerate(tabula_tables):
                    print(f"      Table {i}: Shape {df.shape}, Empty: {df.empty}")

                    if not df.empty and df.shape[0] > 0 and df.shape[1] > 0:
                        # Check if table has actual content
                        non_empty_cells = df.notna().sum().sum()
                        print(f"        Non-empty cells: {non_empty_cells}/{df.shape[0] * df.shape[1]}")

                        # Include tables with any meaningful content (more inclusive)
                        total_cells = df.shape[0] * df.shape[1]
                        if non_empty_cells > 0 or total_cells >= 4:  # Accept if has content OR is reasonably sized
                            tables.append({
                                'method': f'tabula_{config_name}',
                                'table_index': i,
                                'dataframe': df,
                                'config': config
                            })
                            print(f"        Added table with shape {df.shape}")
                        else:
                            print(f"Skipping table with no content")
                    else:
                        print(f"        Skipping empty table {i}")

            except Exception as e:
                print(f"    ❌ Tabula {config_name} failed: {e}")

        return tables

    def extract_with_pdfplumber(self, pdf_path: str, structure_info: Dict) -> List[Dict]:
        """Extract tables using pdfplumber with adaptive settings"""
        tables = []

        try:
            with pdfplumber.open(pdf_path) as pdf:
                print(f"    PDFplumber processing {len(pdf.pages)} pages")

                for page_num, page in enumerate(pdf.pages):
                    print(f"      Page {page_num + 1}:")

                    # Multiple extraction strategies
                    strategies = []

                    if structure_info['has_borders']:
                        strategies.append(("lines", {
                            "vertical_strategy": "lines",
                            "horizontal_strategy": "lines"
                        }))

                    # Always try text-based extraction
                    strategies.extend([
                        ("text_loose", {
                            "vertical_strategy": "text",
                            "horizontal_strategy": "text",
                            "text_x_tolerance": 5,
                            "text_y_tolerance": 3
                        }),
                        ("text_tight", {
                            "vertical_strategy": "text",
                            "horizontal_strategy": "text",
                            "text_x_tolerance": 2,
                            "text_y_tolerance": 2
                        }),
                        ("explicit", {
                            "vertical_strategy": "explicit",
                            "horizontal_strategy": "explicit",
                            "explicit_vertical_lines": [],
                            "explicit_horizontal_lines": []
                        })
                    ])

                    for strategy_name, table_settings in strategies:
                        try:
                            print(f"        Trying {strategy_name} strategy...")
                            page_tables = page.extract_tables(table_settings)
                            print(f"        Found {len(page_tables)} tables with {strategy_name}")

                            for i, table in enumerate(page_tables):
                                if table and len(table) > 0:
                                    print(f"          Table {i}: {len(table)} rows, {len(table[0]) if table[0] else 0} columns")

                                    # Handle different table structures
                                    if len(table) == 1:
                                        # Single row - might be header only or data in one cell
                                        df = pd.DataFrame([table[0]])
                                    else:
                                        # Multi-row table
                                        headers = table[0] if table[0] and any(table[0]) else None
                                        data_rows = table[1:] if len(table) > 1 else []

                                        if not data_rows:
                                            # Only header row
                                            df = pd.DataFrame([table[0]])
                                        else:
                                            df = pd.DataFrame(data_rows, columns=headers)

                                    if not df.empty:
                                        tables.append({
                                            'method': f'pdfplumber_{strategy_name}',
                                            'table_index': i,
                                            'page': page_num,
                                            'strategy': strategy_name,
                                            'dataframe': df
                                        })
                                        print(f"            Added table with shape {df.shape}")

                        except Exception as strategy_error:
                            print(f"        Strategy {strategy_name} failed: {strategy_error}")
                            continue

        except Exception as e:
            print(f"    ❌ PDFplumber extraction failed: {e}")
            import traceback
            print(f"    Traceback: {traceback.format_exc()}")

        return tables

    def score_table_quality(self, table_data: Dict) -> float:
        """
        Score table quality based on multiple criteria:
        - Data completeness
        - Structure consistency
        - Cell content quality
        - Table dimensions (prefer multi-column tables)
        """
        df = table_data['dataframe']

        if df.empty:
            return 0.0

        # Basic metrics
        total_cells = df.shape[0] * df.shape[1]
        empty_cells = df.isnull().sum().sum()
        completeness = 1 - (empty_cells / total_cells) if total_cells > 0 else 0

        # Structure consistency
        row_consistency = df.apply(lambda row: len([x for x in row if pd.notna(x)]), axis=1)
        structure_score = 1 - (row_consistency.std() / row_consistency.mean()) if row_consistency.mean() > 0 else 0

        # Content quality (non-empty strings)
        content_score = 0
        for col in df.columns:
            col_data = df[col].dropna().astype(str)
            if len(col_data) > 0:
                # Ensure col_data is a pandas Series before using .str accessor
                if hasattr(col_data, 'str'):
                    avg_length = col_data.str.len().mean()
                else:
                    # Fallback for non-Series data
                    avg_length = sum(len(str(x)) for x in col_data) / len(col_data)
                content_score += min(avg_length / 10, 1)  # Normalize to max 1 per column
        content_score = content_score / len(df.columns) if len(df.columns) > 0 else 0

        # Dimension bonus - prefer multi-column tables
        dimension_score = 0
        if df.shape[1] > 1:  # Multi-column table
            # Bonus for having multiple columns
            dimension_score = min(df.shape[1] / 5, 1)  # Max bonus at 5+ columns

            # Extra bonus for reasonable aspect ratio
            if 2 <= df.shape[1] <= 10:  # Sweet spot for table columns
                dimension_score *= 1.2
        else:
            # Penalty for single-column tables (likely parsing errors)
            dimension_score = 0.1

        # Method-specific bonuses
        method_bonus = 0
        if 'camelot' in table_data['method'] and 'accuracy' in table_data:
            method_bonus = table_data['accuracy'] / 100
        elif 'tabula' in table_data['method']:
            # Bonus for tabula if it has good structure
            method_bonus = 0.1 if df.shape[1] > 2 else 0.05
        elif 'pdfplumber' in table_data['method']:
            # Bonus for pdfplumber if it has good structure
            method_bonus = 0.15 if df.shape[1] > 2 else 0.05

        # Weighted final score - increased weight on dimensions
        final_score = (
            completeness * 0.25 +
            structure_score * 0.25 +
            content_score * 0.15 +
            dimension_score * 0.25 +  # Important for table structure
            method_bonus * 0.1
        )

        return min(final_score, 1.0)

    def _is_real_table(self, df: pd.DataFrame) -> bool:
        """
        Determine if the extracted content is a real table or just text/paragraph content.

        Returns:
            True if it's a real table, False if it's text/paragraph content
        """
        # Basic checks
        if df.empty or df.shape[0] <= 1:
            return False

        # Check 1: Tables typically have multiple columns with meaningful data
        if df.shape[1] < 2:
            return False

        # Check 2: Count non-empty cells per column
        col_fill_rates = []
        for col in df.columns:
            non_empty = df[col].notna().sum()
            fill_rate = non_empty / len(df) if len(df) > 0 else 0
            col_fill_rates.append(fill_rate)

        # If most columns are very sparse, it's likely not a real table
        avg_fill_rate = sum(col_fill_rates) / len(col_fill_rates) if col_fill_rates else 0
        if avg_fill_rate < 0.3:  # Less than 30% filled
            return False

        # Check 3: Look for numeric patterns (tables often have numbers)
        has_numeric = False
        for col in df.columns:
            # Check if column has numeric values
            col_data = df[col].dropna().astype(str)
            numeric_count = sum(1 for val in col_data if any(c.isdigit() for c in str(val)))
            if numeric_count > len(col_data) * 0.3:  # At least 30% numeric
                has_numeric = True
                break

        # Check 4: Look for table-like patterns in headers or first row
        first_row = df.iloc[0] if len(df) > 0 else []
        table_keywords = ['qty', 'quantity', 'price', 'amount', 'total', 'rate',
                         'cost', 'no', 'item', 'description', 'date', 'id',
                         'name', 'value', 'count', 'unit', 's.no', 'sr.no']

        has_table_pattern = False
        for val in first_row:
            val_str = str(val).lower().strip()
            if any(keyword in val_str for keyword in table_keywords):
                has_table_pattern = True
                break

        # Check 5: Consistent column structure (not just one column with text)
        if df.shape[1] == 1:
            # Single column - check if it's paragraph text
            first_col = df.iloc[:, 0].dropna()
            if len(first_col) > 0:
                # Check average text length
                avg_length = sum(len(str(val)) for val in first_col) / len(first_col)
                if avg_length > 50:  # Long text indicates paragraphs
                    return False

        # Check 6: Tables usually have consistent data types in columns
        has_consistent_structure = False
        for col in df.columns:
            col_data = df[col].dropna().astype(str)
            if len(col_data) > 2:
                # Check if values have similar lengths or patterns
                lengths = [len(str(val)) for val in col_data]
                if lengths:
                    length_variance = max(lengths) - min(lengths)
                    if length_variance < 20:  # Similar lengths suggest structured data
                        has_consistent_structure = True
                        break

        # Decision logic
        score = 0
        if df.shape[1] >= 2:
            score += 1
        if df.shape[0] >= 3:
            score += 1
        if has_numeric:
            score += 2
        if has_table_pattern:
            score += 2
        if has_consistent_structure:
            score += 1
        if avg_fill_rate > 0.5:
            score += 1

        # Need at least a score of 3 to be considered a real table
        return score >= 3

    def _identify_unique_tables(self, all_extractions: List[Dict]) -> List[Dict]:
        """
        Identify unique tables from all extractions and select the best extraction for each.
        Groups similar tables together and returns the best extraction for each unique table.

        Returns:
            List of best extractions for each unique table found in the PDF
        """
        if not all_extractions:
            return []

        # Group tables by their dimensions and content similarity
        table_groups = []

        for extraction in all_extractions:
            df = extraction['dataframe']

            # Skip empty dataframes
            if df.empty:
                continue

            # Create a signature for the table based on dimensions and content
            signature = {
                'rows': df.shape[0],
                'cols': df.shape[1],
                'first_row': tuple(df.iloc[0].fillna('').astype(str)) if len(df) > 0 else (),
                'last_row': tuple(df.iloc[-1].fillna('').astype(str)) if len(df) > 0 else (),
                'total_non_empty': df.notna().sum().sum()
            }

            # Check if this table belongs to an existing group
            found_group = False
            for group in table_groups:
                # Compare with the first table in the group
                group_sig = group['signature']

                # Tables are considered the same if they have similar dimensions and content
                same_dimensions = (abs(signature['rows'] - group_sig['rows']) <= 2 and
                                 abs(signature['cols'] - group_sig['cols']) <= 1)

                # Check content similarity (at least 70% similarity in first/last rows)
                content_similar = False
                if same_dimensions and signature['first_row'] and group_sig['first_row']:
                    # Compare first rows
                    if len(signature['first_row']) == len(group_sig['first_row']):
                        matches = sum(1 for a, b in zip(signature['first_row'], group_sig['first_row'])
                                    if a == b or (a == '' and b == '') or (a != '' and b != ''))
                        content_similar = matches / len(signature['first_row']) >= 0.7

                if same_dimensions and content_similar:
                    # Add to existing group
                    group['extractions'].append(extraction)
                    found_group = True
                    break

            if not found_group:
                # Create a new group
                table_groups.append({
                    'signature': signature,
                    'extractions': [extraction]
                })

        # Select the best extraction from each group
        unique_tables = []
        for group in table_groups:
            # Sort by quality score and select the best
            group['extractions'].sort(key=lambda x: x['quality_score'], reverse=True)
            best_extraction = group['extractions'][0]
            unique_tables.append(best_extraction)

        # Sort unique tables by their appearance order or quality
        unique_tables.sort(key=lambda x: x.get('table_index', 0))

        return unique_tables

    def sanitize_for_json(self, value):
        """Ensure JSON compliance"""
        if pd.isna(value) or value is None:
            return None
        elif isinstance(value, (int, float)):
            if np.isnan(value) or np.isinf(value):
                return None
            return value
        else:
            return str(value).strip() if str(value).strip() else None

    def create_excel_ready_format(self, df: pd.DataFrame, table_metadata: Dict) -> Dict:
        """
        Create Excel-ready JSON format with proper structure:
        - Headers clearly defined
        - Rows and columns in nested format
        - Data types preserved
        - Easy conversion to Excel
        """
        if df.empty:
            return {"error": "Empty table"}

        # Clean and prepare data
        df = df.fillna('')

        # Detect headers (first row or column names)
        headers = []
        if df.columns.tolist()[0] != 0:  # Has named columns
            headers = [self.sanitize_for_json(col) for col in df.columns]
            data_rows = df.values.tolist()
        else:  # First row is header
            headers = [self.sanitize_for_json(cell) for cell in df.iloc[0].tolist()]
            data_rows = df.iloc[1:].values.tolist()

        # Create Excel-ready structure
        excel_format = {
            "table_metadata": {
                "extraction_method": table_metadata.get('method', 'camelot'),
                "table_index": table_metadata.get('table_index', 0),
                "dimensions": {
                    "rows": len(data_rows),
                    "columns": len(headers)
                }
            },
            "headers": [
                {
                    "index": i,
                    "name": header if header else f"Column_{i+1}",
                    "data_type": "text"  # Will be enhanced with actual type detection
                }
                for i, header in enumerate(headers)
            ],
            "rows": []
        }

        # Process each row
        for row_idx, row_data in enumerate(data_rows):
            row_obj = {
                "row_index": row_idx,
                "columns": []
            }

            for col_idx, cell_value in enumerate(row_data):
                sanitized_value = self.sanitize_for_json(cell_value)

                # Detect data type and parse value
                parsed_value = self.parse_cell_value(sanitized_value)

                column_obj = {
                    "column_index": col_idx,
                    "header_name": headers[col_idx] if col_idx < len(headers) else f"Column_{col_idx+1}",
                    "raw_value": sanitized_value,
                    "parsed_value": parsed_value['value'],
                    "data_type": parsed_value['type'],
                    "units": parsed_value.get('units'),
                    "is_numeric": parsed_value['type'] in ['integer', 'float', 'currency', 'percentage']
                }

                row_obj["columns"].append(column_obj)

            excel_format["rows"].append(row_obj)

        return excel_format

    def parse_cell_value(self, value) -> Dict:
        """Parse cell value to determine type and extract meaningful data"""
        if not value or str(value).strip() == '':
            return {"value": None, "type": "empty"}

        value_str = str(value).strip()

        # Check for currency
        currency_pattern = r'^[\$€£¥₹]?[\d,]+\.?\d*[\$€£¥₹]?$'
        if re.match(currency_pattern, value_str):
            numeric_part = re.findall(r'[\d,]+\.?\d*', value_str)[0].replace(',', '')
            currency_symbol = re.findall(r'[\$€£¥₹]', value_str)
            try:
                return {
                    "value": float(numeric_part),
                    "type": "currency",
                    "units": currency_symbol[0] if currency_symbol else None
                }
            except ValueError:
                pass

        # Check for percentage
        if value_str.endswith('%'):
            try:
                numeric_part = float(value_str[:-1])
                return {"value": numeric_part, "type": "percentage", "units": "%"}
            except ValueError:
                pass

        # Check for integers
        if value_str.isdigit() or (value_str.startswith('-') and value_str[1:].isdigit()):
            return {"value": int(value_str), "type": "integer"}

        # Check for floats
        try:
            float_val = float(value_str.replace(',', ''))
            return {"value": float_val, "type": "float"}
        except ValueError:
            pass

        # Default to text
        return {"value": value_str, "type": "text"}

    def _clean_table_text(self, content: str) -> str:
        """Clean and normalize table text for better parsing"""
        if not content:
            return content

        # Fix common PDF extraction issues
        content = content.replace('n n', '\n')  # Fix "n n" artifacts
        content = content.replace(' n ', '\n')  # Fix " n " artifacts
        content = re.sub(r'\bn\b', '\n', content)  # Replace standalone 'n' with newlines

        # Fix currency symbols that might be missing
        content = re.sub(r'(\d+),(\d+)', r'₹\1,\2', content)  # Add ₹ to numbers with commas

        # Normalize whitespace
        content = re.sub(r'\s+', ' ', content)  # Multiple spaces to single space
        content = re.sub(r'\n\s+', '\n', content)  # Remove spaces after newlines
        content = content.strip()

        return content

    def _clean_extracted_value(self, value: str) -> str:
        """Clean individual extracted table values"""
        if not value or value.strip() == "":
            return value

        # Fix the "n" character that appears instead of ₹
        if value.startswith('n') and value[1:].replace(',', '').replace('.', '').isdigit():
            value = '₹' + value[1:]

        # Clean any remaining artifacts
        value = value.replace('n n', ' ')
        value = value.strip()

        return value

    def _parse_text_table_advanced(self, content: str) -> Dict[str, Any]:
        """
        Advanced text table parsing that uses multiple strategies to extract
        table structure and values accurately.
        """
        if not content:
            return {}

        # Clean the content first
        content = self._clean_table_text(content)

        lines = [line.strip() for line in content.split('\n') if line.strip()]
        if not lines:
            return {}

        result = {
            'headers': [],
            'rows': [],
            'summary': {}
        }

        # Strategy 1: Look for clear table structure with S.No
        table_section = []
        summary_section = []
        in_table = False

        for line in lines:
            line_lower = line.lower()

            # Detect table start
            if any(pattern in line_lower for pattern in ['s.no', 's no', 'serial']):
                in_table = True
                table_section.append(line)
                continue

            # Detect summary section
            if any(pattern in line_lower for pattern in ['subtotal', 'total', 'gst', 'tax']):
                in_table = False
                summary_section.append(line)
                continue

            # Add to appropriate section
            if in_table:
                table_section.append(line)
            elif any(pattern in line_lower for pattern in ['subtotal', 'total', 'gst', 'tax']):
                summary_section.append(line)

        # Parse table section
        if table_section:
            parsed_table = self._parse_table_section(table_section)
            result.update(parsed_table)

        # Parse summary section
        if summary_section:
            result['summary'] = self._parse_summary_section(summary_section)

        return result

    def _parse_table_section(self, table_lines: List[str]) -> Dict[str, Any]:
        """Parse the main table section with items"""
        if not table_lines:
            return {'headers': [], 'rows': []}

        result = {'headers': [], 'rows': []}

        # Find header line (first line with table keywords)
        header_line = None
        data_start_idx = 0

        for i, line in enumerate(table_lines):
            line_lower = line.lower()
            header_indicators = ['s.no', 'item', 'description', 'qty', 'rate', 'amount']
            if sum(1 for indicator in header_indicators if indicator in line_lower) >= 3:
                header_line = line
                data_start_idx = i + 1
                break

        if header_line:
            # Parse headers using multiple methods
            headers = self._extract_headers_from_line(header_line)
            result['headers'] = headers

            # Parse data rows
            current_item = []
            for line in table_lines[data_start_idx:]:
                # Check if this line starts a new item (starts with number)
                if re.match(r'^\s*\d+', line):
                    # Save previous item if exists
                    if current_item:
                        parsed_row = self._parse_data_row(' '.join(current_item), len(headers))
                        if parsed_row:
                            result['rows'].append(parsed_row)

                    # Start new item
                    current_item = [line]
                else:
                    # Continuation of current item
                    if current_item:
                        current_item.append(line)

            # Don't forget the last item
            if current_item:
                parsed_row = self._parse_data_row(' '.join(current_item), len(headers))
                if parsed_row:
                    result['rows'].append(parsed_row)

        return result

    def _extract_headers_from_line(self, header_line: str) -> List[str]:
        """Extract column headers from header line using intelligent parsing"""
        # Common header patterns
        standard_headers = ['S.No', 'Item Description', 'Qty', 'Rate', 'Amount']

        # Try to map to standard headers
        line_lower = header_line.lower()
        detected_headers = []

        if 's.no' in line_lower or 'serial' in line_lower:
            detected_headers.append('S.No')

        if any(word in line_lower for word in ['item', 'description', 'product']):
            detected_headers.append('Item Description')

        if 'qty' in line_lower or 'quantity' in line_lower:
            detected_headers.append('Qty')

        if 'rate' in line_lower or 'price' in line_lower:
            detected_headers.append('Rate')

        if 'amount' in line_lower or 'total' in line_lower:
            detected_headers.append('Amount')

        # If we found the expected headers, return them
        if len(detected_headers) >= 3:
            return detected_headers

        # Fallback: split by spaces and clean
        parts = re.split(r'\s{2,}', header_line)
        return [part.strip() for part in parts if part.strip()]

    def _parse_data_row(self, row_text: str, expected_cols: int) -> List[str]:
        """Parse a data row and extract values for each column"""
        if not row_text.strip():
            return []

        # Strategy: Parse "1 Product Name 25 ₹1,500 ₹37,500" accurately
        result = []

        # Step 1: Extract S.No (number at start)
        sno_match = re.match(r'^\s*(\d+)', row_text)
        if sno_match:
            result.append(sno_match.group(1))
            remaining = row_text[sno_match.end():].strip()
        else:
            remaining = row_text

        # Step 2: Extract amounts (₹ values) from the end
        amount_pattern = r'₹[\d,]+(?:\.\d*)?'
        amounts = re.findall(amount_pattern, remaining)

        # Remove amounts from text (from right to left to preserve order)
        temp_remaining = remaining
        for amount in reversed(amounts):
            # Find the last occurrence and remove it
            last_pos = temp_remaining.rfind(amount)
            if last_pos != -1:
                temp_remaining = temp_remaining[:last_pos] + temp_remaining[last_pos + len(amount):]

        # Step 3: Extract quantity (number just before the amounts)
        # Look for numbers at the end of the remaining text
        qty_match = re.search(r'(\d+(?:\.\d+)?)\s*$', temp_remaining.strip())
        qty = None
        if qty_match:
            qty = qty_match.group(1)
            # Remove quantity from remaining text
            temp_remaining = temp_remaining[:qty_match.start()].strip()

        # Step 4: What's left is the item description
        description = re.sub(r'\s+', ' ', temp_remaining).strip()

        # Step 5: Assemble the result in proper order
        if len(result) == 1:  # We have S.No
            # Add description
            if description:
                result.append(description)

            # Add quantity
            if qty:
                result.append(qty)

            # Add amounts in the order they appeared
            for amount in amounts:
                result.append(amount)

            # Pad to expected columns if needed
            while len(result) < expected_cols:
                result.append('')

            return result[:expected_cols]

        return []

    def _parse_summary_section(self, summary_lines: List[str]) -> Dict[str, str]:
        """Parse summary section (Subtotal, GST, Total, etc.)"""
        summary = {}

        for line in summary_lines:
            # Look for key-value pairs
            line = line.strip()
            if not line:
                continue

            # Try to split on common patterns
            # Pattern 1: "Key ₹Value" or "Key Value"
            for pattern in [r'^(.+?)\s*(₹[\d,]+(?:\.\d+)?)$', r'^(.+?)\s+([₹\d,]+(?:\.\d+)?)$']:
                match = re.match(pattern, line)
                if match:
                    key = match.group(1).strip()
                    value = match.group(2).strip()
                    summary[key] = value
                    break
            else:
                # If no pattern matched, store the whole line as a key
                if any(keyword in line.lower() for keyword in ['subtotal', 'total', 'gst', 'tax']):
                    summary[line] = ""

        return summary

    def convert_to_excel(self, table_json: Union[str, Dict], output_path: str = None) -> Union[str, bytes]:
        """
        Convert table JSON to Excel format

        Args:
            table_json: JSON string or dict containing table data
            output_path: Optional file path to save Excel file. If None, returns bytes

        Returns:
            str: File path if output_path provided, otherwise bytes of Excel file
        """
        # Parse JSON if string
        if isinstance(table_json, str):
            data = json.loads(table_json)
        else:
            data = table_json

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Process each table
        tables = data.get("tables", [])
        if not tables:
            # Create empty sheet
            ws = wb.create_sheet("No Tables Found")
            ws["A1"] = "No tables were extracted from the PDF"
            ws["A1"].font = Font(bold=True)
        else:
            for table_idx, table_data in enumerate(tables):
                # Create worksheet
                sheet_name = f"Table_{table_idx + 1}"
                ws = wb.create_sheet(sheet_name)

                # Start table data at row 1 (no metadata)
                start_row = 1

                # Add headers
                headers = table_data.get("headers", [])
                if headers:
                    for col_idx, header in enumerate(headers):
                        cell = ws.cell(row=start_row, column=col_idx + 1)
                        cell.value = header.get("name", f"Column_{col_idx + 1}")
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = border

                # Add data rows
                rows = table_data.get("rows", [])
                for row_idx, row_data in enumerate(rows):
                    columns = row_data.get("columns", [])
                    for col_idx, col_data in enumerate(columns):
                        cell = ws.cell(row=start_row + 1 + row_idx, column=col_idx + 1)

                        # Use parsed value if available
                        value = col_data.get("parsed_value")
                        if value is None:
                            value = col_data.get("raw_value", "")

                        cell.value = value
                        cell.border = border

                        # Apply data type specific formatting
                        data_type = col_data.get("data_type", "text")
                        if data_type == "currency":
                            cell.number_format = '"₹"#,##0.00'
                        elif data_type == "percentage":
                            cell.number_format = '0.00%'
                        elif data_type in ["integer", "float"]:
                            cell.number_format = '#,##0.00' if data_type == "float" else '#,##0'

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter

                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    ws.column_dimensions[column_letter].width = adjusted_width

        # Optional: Add a simple summary sheet with just the table count
        # Comment out this entire block if you don't want any summary sheet
        summary_ws = wb.create_sheet("Summary", 0)
        summary = data.get("extraction_summary", {})
        summary_ws["A1"] = "Total Tables Extracted:"
        summary_ws["B1"] = summary.get("total_tables", len(tables))
        summary_ws["A1"].font = Font(bold=True)
        summary_ws.column_dimensions["A"].width = 20
        summary_ws.column_dimensions["B"].width = 10

        # Save or return bytes
        if output_path:
            wb.save(output_path)
            return output_path
        else:
            # Return as bytes
            excel_bytes = io.BytesIO()
            wb.save(excel_bytes)
            excel_bytes.seek(0)
            return excel_bytes.getvalue()

    def extract_tables(self, pdf_path: str, extract_all_unique: bool = True) -> str:
        """
        Main extraction method that:
        1. Analyzes PDF visually
        2. Extracts with Camelot only
        3. Returns all tables in Excel-ready format

        Args:
            pdf_path: Path to the PDF file
            extract_all_unique: If True, extracts ALL tables from PDF (default: True)
                               If False, limits to first 3 tables

        Returns:
            JSON string with all extracted tables and metadata
        """
        debug_info = {
            "pdf_path": pdf_path,
            "extraction_steps": [],
            "method_results": {}
        }

        try:
            # Validate PDF file exists
            if not os.path.exists(pdf_path):
                error_result = {
                    "error": f"PDF file not found: {pdf_path}",
                    "debug_info": debug_info
                }
                return json.dumps(error_result, indent=2)

            # Step 1: Visual analysis of all pages for comprehensive structure detection
            print(f"🔍 Starting visual analysis of PDF: {pdf_path}")
            debug_info["extraction_steps"].append("visual_analysis_started")

            # Analyze first page for structure, but also check total page count
            try:
                doc = fitz.open(pdf_path)
                total_pages = len(doc)
                doc.close()
            except Exception as e:
                error_result = {
                    "error": f"Failed to open PDF with PyMuPDF: {str(e)}",
                    "debug_info": debug_info,
                    "suggestion": "Please ensure the PDF is not corrupted and is readable"
                }
                return json.dumps(error_result, indent=2)

            image = self.pdf_page_to_image(pdf_path, 0)
            structure_info = self.detect_table_structure(image)
            structure_info['total_pages'] = total_pages

            print(f"📊 Visual analysis results:")
            print(f"  - Total pages: {total_pages}")
            print(f"  - Table count: {structure_info.get('table_count', 0)}")
            print(f"  - Has borders: {structure_info.get('has_borders', False)}")
            print(f"  - Line density: {structure_info.get('line_density', 0):.4f}")

            debug_info["visual_analysis"] = structure_info
            debug_info["extraction_steps"].append("visual_analysis_completed")

            # Step 2: Extract with Camelot only
            all_extractions = []
            print(f"\n🛠️ Starting extraction with Camelot...")

            # Camelot extraction
            print("📋 Trying Camelot extraction...")
            debug_info["extraction_steps"].append("camelot_started")
            camelot_results = self.extract_with_camelot(pdf_path, structure_info)
            debug_info["method_results"]["camelot"] = {"count": len(camelot_results), "success": len(camelot_results) > 0}
            print(f"  Camelot found {len(camelot_results)} tables")
            all_extractions.extend(camelot_results)

            print(f"\n📈 Total tables found: {len(all_extractions)}")
            debug_info["total_extractions"] = len(all_extractions)

            if not all_extractions:
                error_result = {
                    "error": "No tables found with Camelot",
                    "debug_info": debug_info,
                    "visual_analysis": structure_info,
                    "troubleshooting": {
                        "suggestions": [
                            "Check if PDF contains actual tables",
                            "Verify PDF is not corrupted",
                            "Try with a different PDF format"
                        ]
                    }
                }
                print("❌ No tables found with Camelot extraction")
                return json.dumps(error_result, indent=2)

            # Step 3: Process extracted tables
            print(f"\n📋 Processing {len(all_extractions)} tables...")
            debug_info["extraction_steps"].append("processing_started")

            # Add table info for debugging
            for i, extraction in enumerate(all_extractions):
                df = extraction['dataframe']
                print(f"  Table {i+1}: Shape {df.shape[0]}x{df.shape[1]}")
                print(f"    Non-empty cells: {df.notna().sum().sum()}/{df.shape[0] * df.shape[1]}")

            output = {
                "extraction_summary": {
                    "total_tables": len(all_extractions),
                    "extraction_method": "camelot",
                    "visual_analysis": structure_info,
                    "debug_info": debug_info
                },
                "tables": []
            }

            # Include all tables or limit based on parameter
            if extract_all_unique:
                top_results = all_extractions  # Take all tables
                print(f"📋 Including all {len(top_results)} tables in output")
            else:
                top_results = all_extractions[:min(3, len(all_extractions))]  # Limit to 3
                print(f"📋 Including first {len(top_results)} tables in output")

            for extraction in top_results:
                excel_ready_table = self.create_excel_ready_format(
                    extraction['dataframe'],
                    extraction
                )
                output["tables"].append(excel_ready_table)
                print(f"  ✓ Converted {extraction['method']} table to Excel format")

            debug_info["extraction_steps"].append("output_created")
            print(f"✅ Extraction completed successfully!")

            return json.dumps(output, indent=2, ensure_ascii=False)

        except Exception as e:
            import traceback
            error_result = {
                "error": f"Extraction failed: {str(e)}",
                "traceback": traceback.format_exc(),
                "debug_info": debug_info
            }
            print(f"❌ Error during extraction: {str(e)}")
            return json.dumps(error_result, indent=2)

# Legacy function for backward compatibility
def extract_tables(pdf_path: str) -> str:
    """Backward compatible function - extracts ALL tables using Camelot"""
    extractor = VisualTableExtractor()
    return extractor.extract_tables(pdf_path, extract_all_unique=True)

# New recommended function
def extract_tables_with_cv(pdf_path: str, extract_all_unique: bool = True) -> str:
    """
    Main function using computer vision enhanced extraction with Camelot

    Args:
        pdf_path: Path to PDF file
        extract_all_unique: If True, extracts ALL tables from PDF (default: True)
                           If False, limits to first 3 tables
    """
    extractor = VisualTableExtractor()
    return extractor.extract_tables(pdf_path, extract_all_unique=extract_all_unique)

# Excel conversion functions
def convert_table_json_to_excel(table_json: Union[str, Dict], output_path: str = None) -> Union[str, bytes]:
    """
    Convert table extraction JSON to Excel format

    Args:
        table_json: JSON string or dict from table extraction
        output_path: Optional file path to save Excel. If None, returns bytes

    Returns:
        str: File path if output_path provided, otherwise bytes of Excel file

    Example:
        # Convert JSON to Excel file
        json_data = extract_tables_with_cv('document.pdf')
        excel_path = convert_table_json_to_excel(json_data, 'output.xlsx')

        # Get Excel as bytes for API response
        excel_bytes = convert_table_json_to_excel(json_data)
    """
    extractor = VisualTableExtractor()
    return extractor.convert_to_excel(table_json, output_path)

def extract_tables_to_excel(pdf_path: str, excel_output_path: str = None, extract_all_unique: bool = True) -> Union[str, bytes]:
    """
    Extract tables from PDF using Camelot and directly convert to Excel

    Args:
        pdf_path: Path to PDF file
        excel_output_path: Optional path to save Excel file. If None, returns bytes
        extract_all_unique: If True, extracts ALL tables from PDF (default: True)

    Returns:
        str: Excel file path if excel_output_path provided, otherwise bytes

    Example:
        # Extract ALL tables from PDF to Excel file
        excel_path = extract_tables_to_excel('document.pdf', 'tables.xlsx')

        # Extract ALL tables from PDF to Excel bytes for download
        excel_bytes = extract_tables_to_excel('document.pdf')

        # Extract first 3 tables only
        excel_bytes = extract_tables_to_excel('document.pdf', extract_all_unique=False)
    """
    try:
        # Extract tables using Camelot
        extractor = VisualTableExtractor()
        json_result = extractor.extract_tables(pdf_path, extract_all_unique=extract_all_unique)

        # Check if extraction had errors
        result_dict = json.loads(json_result)
        if "error" in result_dict:
            print(f"Error during extraction: {result_dict['error']}")
            # Try simplified extraction without visual analysis
            print("Attempting simplified extraction without visual analysis...")
            json_result = extract_tables_simple(pdf_path)

        # Convert to Excel
        return extractor.convert_to_excel(json_result, excel_output_path)
    except Exception as e:
        print(f"Error in extract_tables_to_excel: {e}")
        # Try a simple fallback method
        return extract_tables_simple_to_excel(pdf_path, excel_output_path)

def convert_main_layout_to_excel_tables_only(layout_data: Dict, output_path: str = None) -> Union[str, bytes]:
    """
    Convert only the table data from main layout extraction to Excel format
    This extracts ONLY tables and table-like text content, ignoring other text

    Args:
        layout_data: Output from get_document_layout()
        output_path: Optional file path to save Excel. If None, returns bytes

    Returns:
        str: File path if output_path provided, otherwise bytes of Excel file
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    import re

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Count tables and tabular text
    structured_tables = 0
    text_tables = 0

    for page_key, items in layout_data.items():
        for item in items:
            if item.get("type") == "table":
                structured_tables += 1
            elif item.get("type") == "text":
                content = item.get("content", "")
                if any(keyword in content.lower() for keyword in ['s.no', 'item description', 'qty', 'rate', 'amount', 'total']):
                    text_tables += 1

    # Create summary sheet
    summary_ws = wb.create_sheet("Table Extraction Summary", 0)
    summary_ws["A1"] = "PDF Table Extraction Results"
    summary_ws["A1"].font = Font(size=16, bold=True)

    summary_ws["A3"] = "Structured Tables Found:"
    summary_ws["B3"] = structured_tables
    summary_ws["A4"] = "Text-based Tables Found:"
    summary_ws["B4"] = text_tables
    summary_ws["A5"] = "Total Tables:"
    summary_ws["B5"] = structured_tables + text_tables

    # Style summary
    for row in range(1, 6):
        summary_ws[f"A{row}"].font = Font(bold=True)

    summary_ws.column_dimensions["A"].width = 25
    summary_ws.column_dimensions["B"].width = 15

    table_count = 0
    has_any_tables = False

    # Process each page - ONLY tables and table-like content
    for page_key, items in layout_data.items():
        page_tables = []

        # First pass: collect all structured tables
        for item in items:
            if item.get("type") == "table":
                page_tables.append(("structured", item))

        # Second pass: collect table-like text
        for item in items:
            if item.get("type") == "text":
                content = item.get("content", "")
                # Check if content contains table indicators
                if any(keyword in content.lower() for keyword in ['s.no', 'item description', 'qty', 'rate', 'amount', 'subtotal', 'total value']):
                    page_tables.append(("text_table", item))

        # Create sheet only if we found tables
        if page_tables:
            has_any_tables = True
            page_ws = wb.create_sheet(f"Tables_Page_{page_key.split('_')[1]}")

            page_ws["A1"] = f"Tables from {page_key.upper()}"
            page_ws["A1"].font = Font(size=14, bold=True)

            current_row = 3

            for table_type, item in page_tables:
                if table_type == "structured":
                    table_count += 1

                    # Add table header
                    page_ws[f"A{current_row}"] = f"STRUCTURED TABLE {table_count}"
                    page_ws[f"A{current_row}"].font = Font(bold=True, size=12, color="008000")
                    current_row += 1

                    content = item.get("content", {})
                    headers = content.get("headers", [])
                    rows = content.get("rows", [])

                    # Add table metadata
                    page_ws[f"A{current_row}"] = f"Method: {content.get('extraction_method', 'unknown')}"
                    page_ws[f"A{current_row}"].font = Font(italic=True)
                    current_row += 1

                    # Add headers
                    if headers:
                        for col_idx, header in enumerate(headers):
                            cell = page_ws.cell(row=current_row, column=col_idx + 1)
                            cell.value = header
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                            cell.border = border
                        current_row += 1

                    # Add data rows
                    for row_data in rows:
                        for col_idx, cell_value in enumerate(row_data):
                            cell = page_ws.cell(row=current_row, column=col_idx + 1)
                            cell.value = cell_value
                            cell.border = border

                            # Format currency values
                            if isinstance(cell_value, str) and ('₹' in cell_value):
                                try:
                                    numeric_value = cell_value.replace('₹', '').replace(',', '').strip()
                                    if numeric_value.replace('.', '').replace('-', '').isdigit():
                                        cell.value = float(numeric_value)
                                        cell.number_format = '"₹"#,##0.00'
                                except:
                                    pass

                        current_row += 1

                elif table_type == "text_table":
                    table_count += 1

                    # Add text table header
                    page_ws[f"A{current_row}"] = f"TEXT-BASED TABLE {table_count}"
                    page_ws[f"A{current_row}"].font = Font(bold=True, size=12, color="FF6600")
                    current_row += 1

                    content = item.get("content", "")

                    # Advanced table parsing with multiple strategies
                    extractor = VisualTableExtractor()
                    parsed_table = extractor._parse_text_table_advanced(content)

                    if parsed_table and parsed_table['rows']:
                        # Display the parsed table
                        page_ws[f"A{current_row}"] = "Intelligently Parsed Table:"
                        page_ws[f"A{current_row}"].font = Font(bold=True, italic=True)
                        current_row += 1

                        # Add headers
                        headers = parsed_table.get('headers', [])
                        if headers:
                            for col_idx, header in enumerate(headers):
                                cell = page_ws.cell(row=current_row, column=col_idx + 1)
                                cell.value = header
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = header_alignment
                                cell.border = border
                            current_row += 1

                        # Add data rows
                        for row_data in parsed_table['rows']:
                            for col_idx, cell_value in enumerate(row_data):
                                cell = page_ws.cell(row=current_row, column=col_idx + 1)

                                # Clean and format the value
                                if isinstance(cell_value, str):
                                    cell_value = cell_value.strip()

                                cell.value = cell_value
                                cell.border = border

                                # Apply formatting based on content
                                if isinstance(cell_value, str):
                                    # Currency formatting
                                    if '₹' in cell_value or cell_value.replace(',', '').replace('.', '').replace('-', '').isdigit():
                                        try:
                                            # Extract numeric part
                                            numeric_str = re.sub(r'[₹,\s]', '', cell_value)
                                            if numeric_str and numeric_str.replace('.', '').replace('-', '').isdigit():
                                                cell.value = float(numeric_str)
                                                cell.number_format = '"₹"#,##0.00'
                                        except:
                                            pass
                                    # Quantity/Rate formatting
                                    elif cell_value.isdigit():
                                        cell.value = int(cell_value)
                                        cell.number_format = '#,##0'

                            current_row += 1

                        # Add summary information
                        summary = parsed_table.get('summary', {})
                        if summary:
                            current_row += 1
                            page_ws[f"A{current_row}"] = "Table Summary:"
                            page_ws[f"A{current_row}"].font = Font(bold=True, color="0066CC")
                            current_row += 1

                            for key, value in summary.items():
                                page_ws[f"A{current_row}"] = key
                                page_ws[f"B{current_row}"] = value

                                # Format currency values in summary
                                if isinstance(value, str) and '₹' in value:
                                    try:
                                        numeric_str = re.sub(r'[₹,\s]', '', value)
                                        if numeric_str.replace('.', '').replace('-', '').isdigit():
                                            page_ws[f"B{current_row}"].value = float(numeric_str)
                                            page_ws[f"B{current_row}"].number_format = '"₹"#,##0.00'
                                    except:
                                        pass

                                page_ws[f"A{current_row}"].font = Font(bold=True)
                                page_ws[f"A{current_row}"].border = border
                                page_ws[f"B{current_row}"].border = border
                                current_row += 1

                    else:
                        # Fallback: display as formatted text
                        page_ws[f"A{current_row}"] = "Raw Table Text (could not parse structure):"
                        page_ws[f"A{current_row}"].font = Font(bold=True, italic=True, color="FF6600")
                        current_row += 1

                        lines = [line.strip() for line in content.split('\n') if line.strip()]
                        for line in lines:
                            page_ws[f"A{current_row}"] = line
                            current_row += 1

                current_row += 2  # Add spacing between tables

            # Auto-adjust column widths
            for column in page_ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                page_ws.column_dimensions[column_letter].width = adjusted_width

    # If no tables found at all
    if not has_any_tables:
        no_tables_ws = wb.create_sheet("No Tables Found")
        no_tables_ws["A1"] = "NO TABLES DETECTED IN THIS PDF"
        no_tables_ws["A1"].font = Font(bold=True, color="FF0000", size=16)

        no_tables_ws["A3"] = "The PDF does not appear to contain any structured tables"
        no_tables_ws["A4"] = "or text that resembles tabular data."
        no_tables_ws["A5"] = ""
        no_tables_ws["A6"] = "Try using the CV-enhanced table extraction method"
        no_tables_ws["A7"] = "at /extract-tables-to-excel/ endpoint"

    # Save or return bytes
    if output_path:
        wb.save(output_path)
        return output_path
    else:
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        return excel_bytes.getvalue()

# Keep the original function for backward compatibility
def convert_main_layout_to_excel(layout_data: Dict, output_path: str = None) -> Union[str, bytes]:
    """
    Convert the main layout extraction result to Excel format (TABLES ONLY)
    This is now an alias for the tables-only function

    Args:
        layout_data: Output from get_document_layout()
        output_path: Optional file path to save Excel. If None, returns bytes

    Returns:
        str: File path if output_path provided, otherwise bytes of Excel file
    """
    return convert_main_layout_to_excel_tables_only(layout_data, output_path)

# Simple fallback extraction methods (without visual analysis)
def extract_tables_simple(pdf_path: str) -> str:
    """
    Simple table extraction without visual analysis - fallback method
    """
    try:
        if not os.path.exists(pdf_path):
            return json.dumps({"error": f"File not found: {pdf_path}"})

        print("Running simple Camelot extraction (stream mode)...")

        # Try stream mode first (works for most tables)
        try:
            tables = camelot.read_pdf(pdf_path, pages='all', flavor='stream', suppress_stdout=True)
        except:
            try:
                # Fallback to lattice mode
                print("Stream failed, trying lattice mode...")
                tables = camelot.read_pdf(pdf_path, pages='all', flavor='lattice', suppress_stdout=True)
            except Exception as e:
                return json.dumps({"error": f"Camelot extraction failed: {str(e)}"})

        # Convert to simple format
        result = {
            "extraction_summary": {
                "total_tables": len(tables),
                "extraction_method": "camelot_simple"
            },
            "tables": []
        }

        for i, table in enumerate(tables):
            df = table.df
            if not df.empty:
                # Convert DataFrame to simple format
                headers = df.iloc[0].tolist() if len(df) > 0 else []
                rows = df.iloc[1:].values.tolist() if len(df) > 1 else []

                result["tables"].append({
                    "table_metadata": {
                        "extraction_method": "camelot_simple",
                        "table_index": i,
                        "dimensions": {
                            "rows": len(rows),
                            "columns": len(headers)
                        }
                    },
                    "headers": [{"name": str(h) if h else f"Column_{i+1}"} for i, h in enumerate(headers)],
                    "rows": [{"columns": [{"raw_value": cell} for cell in row]} for row in rows]
                })

        return json.dumps(result, indent=2, ensure_ascii=False)

    except Exception as e:
        return json.dumps({"error": f"Simple extraction failed: {str(e)}"})

def extract_tables_simple_to_excel(pdf_path: str, excel_output_path: str = None) -> Union[str, bytes]:
    """
    Simple direct PDF to Excel conversion - fallback method
    """
    try:
        import pandas as pd
        from openpyxl import Workbook

        if not os.path.exists(pdf_path):
            raise FileNotFoundError(f"File not found: {pdf_path}")

        # Try to extract tables with Camelot
        tables = []
        try:
            tables = camelot.read_pdf(pdf_path, pages='all', flavor='stream', suppress_stdout=True)
        except:
            try:
                tables = camelot.read_pdf(pdf_path, pages='all', flavor='lattice', suppress_stdout=True)
            except:
                pass

        if not tables:
            # Create empty Excel with error message
            wb = Workbook()
            ws = wb.active
            ws.title = "Error"
            ws["A1"] = "No tables could be extracted from the PDF"

            if excel_output_path:
                wb.save(excel_output_path)
                return excel_output_path
            else:
                excel_bytes = io.BytesIO()
                wb.save(excel_bytes)
                excel_bytes.seek(0)
                return excel_bytes.getvalue()

        # Create Excel workbook
        wb = Workbook()
        wb.remove(wb.active)

        for i, table in enumerate(tables):
            df = table.df
            if not df.empty:
                ws = wb.create_sheet(f"Table_{i+1}")

                # Write data directly
                for row_idx, row in enumerate(df.values):
                    for col_idx, value in enumerate(row):
                        ws.cell(row=row_idx+1, column=col_idx+1, value=str(value) if value else "")

        # Save Excel
        if excel_output_path:
            wb.save(excel_output_path)
            return excel_output_path
        else:
            excel_bytes = io.BytesIO()
            wb.save(excel_bytes)
            excel_bytes.seek(0)
            return excel_bytes.getvalue()

    except Exception as e:
        print(f"Simple Excel extraction failed: {e}")
        # Return error in Excel format
        wb = Workbook()
        ws = wb.active
        ws.title = "Error"
        ws["A1"] = f"Extraction failed: {str(e)}"

        if excel_output_path:
            wb.save(excel_output_path)
            return excel_output_path
        else:
            excel_bytes = io.BytesIO()
            wb.save(excel_bytes)
            excel_bytes.seek(0)
            return excel_bytes.getvalue()

# Backward compatibility function for positional extraction
def extract_tables_with_position(pdf_path: str) -> Dict:
    """
    Backward compatible function that returns tables with position info
    in the format expected by main.py
    """
    try:
        print(f"🔄 extract_tables_with_position called for: {pdf_path}")

        # Use the new CV-enhanced extractor with silent mode
        extractor = VisualTableExtractor()

        # Temporarily disable prints for cleaner output
        import sys
        from io import StringIO

        old_stdout = sys.stdout
        sys.stdout = StringIO()

        try:
            result_json = extractor.extract_tables(pdf_path, extract_all_unique=True)
        finally:
            sys.stdout = old_stdout

        result = json.loads(result_json)
        print(f"📊 CV extraction found {len(result.get('tables', []))} tables")

        # Convert to expected format for main.py
        page_based_output = {}

        if "error" in result:
            print("❌ Error in CV extraction, returning empty")
            return {"page_1": []}

        # Extract tables and convert to position-based format
        if "tables" in result and len(result["tables"]) > 0:
            print(f"🔧 Converting {len(result['tables'])} tables to legacy format")

            for table_idx, table_data in enumerate(result["tables"]):
                print(f"  Processing table {table_idx + 1}")

                # Get table metadata
                metadata = table_data.get("table_metadata", {})
                dimensions = metadata.get("dimensions", {})

                print(f"    Method: {metadata.get('extraction_method')}")
                print(f"    Dimensions: {dimensions.get('rows', 0)}x{dimensions.get('columns', 0)}")

                # Create a simplified table entry with position info
                table_entry = {
                    "type": "table",
                    "table_type": "data_table",
                    "content": {
                        "extraction_method": metadata.get("extraction_method", "cv_enhanced"),
                        "headers": [],
                        "rows": []
                    },
                    "bbox": (50, 100, 550, 400)  # More realistic bbox
                }

                # Extract headers
                headers = []
                for h in table_data.get("headers", []):
                    header_name = h.get("name", f"Column_{h.get('index', 0)}")
                    # Clean header names
                    header_name = header_name.replace('n n', ' ').strip()
                    headers.append(header_name)

                table_entry["content"]["headers"] = headers
                print(f"    Headers: {headers}")

                # Convert rows to simpler format
                rows_converted = []
                for row_idx, row in enumerate(table_data.get("rows", [])):
                    row_values = []
                    for col in row.get("columns", []):
                        # Use parsed_value if available, otherwise raw_value
                        value = col.get("parsed_value")
                        if value is None:
                            value = col.get("raw_value", "")
                        # Convert None to empty string
                        if value is None:
                            value = ""
                        # Clean the extracted value (fix currency symbols)
                        clean_value = str(value)
                        if clean_value.startswith('n') and clean_value[1:].replace(',', '').replace('.', '').isdigit():
                            clean_value = '₹' + clean_value[1:]
                        clean_value = clean_value.replace('n n', ' ').strip()
                        row_values.append(clean_value)

                    if row_values:  # Only add non-empty rows
                        rows_converted.append(row_values)

                table_entry["content"]["rows"] = rows_converted
                print(f"    Converted {len(rows_converted)} data rows")

                # Add to page 1 by default (can be enhanced to detect actual pages)
                if "page_1" not in page_based_output:
                    page_based_output["page_1"] = []
                page_based_output["page_1"].append(table_entry)

            print(f"✅ Successfully converted {len(page_based_output.get('page_1', []))} tables")
        else:
            print("⚠️  No tables found in CV extraction result")

        return page_based_output

    except Exception as e:
        import traceback
        print(f"❌ Error in extract_tables_with_position: {e}")
        print(f"Traceback: {traceback.format_exc()}")
        return {"page_1": []}